# -*- coding: utf-8 -*-
from collections import defaultdict

from odoo import api, fields, models
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import base64

class GeneralLedger(models.TransientModel):
    _name = 'general.ledger'
    _description = 'General Ledger'

    # Definición de campos para el modelo
    report_from_date = fields.Date(string="Reporte desde", required=True, default=fields.Date.context_today)
    report_to_date = fields.Date(string="Reporte hasta", required=True, default=fields.Date.context_today)
    company_id = fields.Many2one('res.company', string="Compañía", default=lambda self: self.env.company)
    file_content = fields.Binary(string="Archivo Contenido")
    file_name = fields.Char(string="Nombre del Archivo", default="Libro Mayor.xlsx")

    def action_generate_excel(self):
        # Crear un libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Definir estilos
        bold_font = Font(bold=True)
        large_font = Font(size=14, bold=True)
        center_alignment = Alignment(horizontal='center')

        # Agregar el nombre de la compañía en mayúsculas, en negrita y centrado
        company_name = self.company_id.name.upper() if self.company_id else "NO DISPONIBLE"
        ws.merge_cells('A1:F1')  # Fusionar celdas de A1 a F1
        ws['A1'] = company_name
        ws['A1'].font = large_font
        ws['A1'].alignment = center_alignment

        # Agregar el título del reporte, en negrita y centrado
        report_title = f"LIBRO MAYOR CORRESPONDIENTE DEL {self.report_from_date} AL {self.report_to_date}"
        ws.merge_cells('A2:F2')  # Fusionar celdas de A2 a F2
        ws['A2'] = report_title
        ws['A2'].font = bold_font
        ws['A2'].alignment = center_alignment

        # Agregar el detalle de la moneda, centrado y en negrita
        currency_detail = "Expresado en: USD"
        ws.merge_cells('A3:F3')  # Fusionar celdas de A3 a F3
        cell = ws['A3']
        cell.value = currency_detail
        cell.font = bold_font  # Aplicar negrita
        cell.alignment = center_alignment

        # Agregar encabezados de la tabla
        headers = ['Codigo', 'Cuenta de mayor', 'Fecha', 'Debe', 'Haber', 'Saldo']
        ws.append(headers)
        for cell in ws[4]:  # Aplicar formato a los encabezados de la tabla
            cell.font = bold_font
            cell.alignment = center_alignment

        # Configurar el ancho de las columnas
        ws.column_dimensions['A'].width = 10  # Columna A
        ws.column_dimensions['B'].width = 25  # Columna B
        ws.column_dimensions['C'].width = 13  # Columna C
        ws.column_dimensions['D'].width = 10  # Columna D
        ws.column_dimensions['E'].width = 10  # Columna E
        ws.column_dimensions['F'].width = 10  # Columna F

        # Filtrar grupos de cuentas por prefijo
        group_model = self.env['account.group']
        groups = group_model.search([('major_account', '=', True)])

        row_index = 7  # Fila de inicio para los datos

        # Inicializar los totales globales
        grand_total_debit = 0.0
        grand_total_credit = 0.0
        grand_total_balance = 0.0

        for group in groups:
            account_prefix = group.code_prefix_start
            accounts = self.env['account.account'].search([('code', 'like', f'{account_prefix}%')])

            # Obtener movimientos contables para las cuentas filtradas en el rango de fechas
            moves = self.env['account.move.line'].search([
                ('account_id', 'in', accounts.ids),
                ('date', '>=', self.report_from_date),
                ('date', '<=', self.report_to_date)
            ])

            # Calcular saldos iniciales consolidados para el grupo de cuentas
            initial_debit = 0.0
            initial_credit = 0.0
            initial_balance = 0.0
            for account in accounts:
                initial_moves = self.env['account.move.line'].search([
                    ('account_id', '=', account.id),
                    ('date', '<', self.report_from_date)
                ])
                for move in initial_moves:
                    initial_debit += move.debit
                    initial_credit += move.credit
                    initial_balance += move.debit - move.credit

            # Calcular los totales históricos (desde la creación de las cuentas hasta la fecha de fin del reporte)
            total_debit = 0.0
            total_credit = 0.0
            total_balance = 0.0
            all_moves = self.env['account.move.line'].search([
                ('account_id', 'in', accounts.ids),
                ('date', '<=', self.report_to_date)
            ])
            for move in all_moves:
                total_debit += move.debit
                total_credit += move.credit
                total_balance += move.debit - move.credit

            # Agregar datos del grupo en la fila 7
            ws.cell(row=row_index, column=1, value=account_prefix).font = bold_font
            ws.cell(row=row_index, column=2, value=group.name).font = bold_font
            ws.cell(row=row_index, column=4, value=total_debit).font = bold_font
            ws.cell(row=row_index, column=5, value=total_credit).font = bold_font
            ws.cell(row=row_index, column=6, value=total_balance).font = bold_font

            row_index += 1

            # Agregar la fila de saldo inicial para Debe, Haber, y Saldo
            ws.append([
                '',  # Código de cuenta en blanco
                'SALDO INICIAL',  # Etiqueta para el saldo inicial
                '',  # Celdas vacías para fecha y código de cuenta
                initial_debit,  # Saldo inicial Debe
                initial_credit,  # Saldo inicial Haber
                initial_balance  # Saldo inicial Saldo
            ])

            # Aplicar formato en negrita a toda la fila de saldo inicial
            for cell in ws[row_index]:
                cell.font = bold_font
            row_index += 1

            # Agrupar movimientos por fecha
            date_summaries = defaultdict(lambda: {'debit': 0.0, 'credit': 0.0})

            for move in moves:
                date_summaries[move.date]['debit'] += move.debit
                date_summaries[move.date]['credit'] += move.credit

            # Ordenar las fechas de menor a mayor
            sorted_dates = sorted(date_summaries.keys())

            # Agregar los datos agrupados por fecha
            accumulated_debit = initial_debit
            accumulated_credit = initial_credit
            accumulated_balance = initial_balance

            for date in sorted_dates:
                summary = date_summaries[date]
                accumulated_debit += summary['debit']
                accumulated_credit += summary['credit']
                accumulated_balance = accumulated_debit - accumulated_credit

                ws.append([
                    "",  # Código de cuenta en blanco
                    "Movimiento del",  # Reemplazado por "Movimiento del"
                    date,
                    summary['debit'],
                    summary['credit'],
                    summary['debit'] - summary['credit']
                ])
                row_index += 1

            # Agregar la fila de suma para cada grupo de cuentas
            ws.append([
                '',  # Código de cuenta en blanco
                'SUMA',  # Etiqueta para la suma
                '',  # Celdas vacías para fecha y código de cuenta
                accumulated_debit,  # Total acumulado Debe
                accumulated_credit,  # Total acumulado Haber
                accumulated_balance  # Total acumulado Saldo
            ])

            # Aplicar formato en negrita a toda la fila de suma
            for cell in ws[row_index]:
                cell.font = bold_font
            row_index += 1

            # Acumulación global de totales
            grand_total_debit += accumulated_debit
            grand_total_credit += accumulated_credit
            grand_total_balance += accumulated_balance

        # Agregar la fila de suma total dos filas después de la última fila de suma de los grupos
        ws.append(['', 'SALDOS TOTALES', '', grand_total_debit, grand_total_credit, grand_total_balance])

        # Aplicar formato en negrita a toda la fila de suma total
        for cell in ws[row_index]:
            cell.font = bold_font
        row_index += 1

        # Guardar el archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Codificar el archivo en base64
        file_content = base64.b64encode(output.read()).decode('utf-8')
        output.close()

        # Asignar el contenido del archivo al campo del modelo
        self.write({
            'file_content': file_content,
            'file_name': "Libro Mayor.xlsx",
        })

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=general.ledger&id=%s&field=file_content&download=true&filename=%s' % (self.id, "Libro Mayor.xlsx"),
            'target': 'self',
        }
